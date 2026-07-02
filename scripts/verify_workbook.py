"""Phase 9 - deterministic verification of the built workbook. Red = do not ship.

Checks:
  1. No already-seen job id leaks in: every exclusion_sources CSV id and every
     master-library id (minus this run's own kept set) must be absent.
  2. Job Applications: No. sequential, links well-formed, status "Not Applied",
     JD non-trivial, sheet id set == kept_candidates id set.
  3. This-Week by Fit: scores strictly non-increasing.
  4. Headers match the v3 spec.

Grew out of the 2026-07-02 adversarial review; runs after every build so a
scoring or dedupe regression can never silently reach the user.
"""
import argparse, csv, datetime, glob, json, os, re, sys
sys.path.insert(0, os.path.dirname(__file__))
import lib_common as L
from openpyxl import load_workbook

SKILL = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MASTER = os.path.join(SKILL, "state", "master_jobs.jsonl")

ap = argparse.ArgumentParser()
ap.add_argument("--config", required=True)
ap.add_argument("--run", required=True)
ap.add_argument("--workbook", default=None, help="explicit xlsx path (default: newest matching basename)")
a = ap.parse_args()
cfg = L.load_config(a.config)

wb_path = a.workbook
if not wb_path:
    pat = os.path.join(cfg["output"]["dir"], cfg["output"].get("basename", "*") + "_*.xlsx")
    matches = sorted(glob.glob(pat), key=os.path.getmtime)
    if not matches:
        print(f"verify_workbook: no workbook matching {pat}")
        sys.exit(1)
    wb_path = matches[-1]

kept_ids = {k["jid"] for k in L.read_json(os.path.join(a.run, "kept_candidates.json"))}

seen_ids = set()
for src in cfg.get("exclusion_sources", []) or []:
    if not os.path.exists(src):
        continue
    for row in csv.DictReader(open(src, encoding="utf-8", errors="replace")):
        g = {k.lstrip("﻿").strip().lower(): v for k, v in row.items()}
        m = re.search(r"(\d{6,})", g.get("url") or "")
        if m:
            seen_ids.add(m.group(1))
if os.path.exists(MASTER):
    for line in open(MASTER, encoding="utf-8", errors="replace"):
        try:
            jid = str(json.loads(line).get("jid"))
        except Exception:
            continue
        if jid and jid not in kept_ids:   # this run's own ids are in master post-build
            seen_ids.add(jid)

wb = load_workbook(wb_path)
ws, ws2 = wb["Job Applications"], wb["This-Week by Fit"]
errors, warns = [], []

ids_in_sheet, prev_no = [], 0
for r in range(2, ws.max_row + 1):
    no = ws.cell(r, 1).value
    link = ws.cell(r, 7)
    url = link.hyperlink.target if link.hyperlink else ""
    m = re.search(r"/jobs/view/(\d+)", url or "")
    if not m:
        errors.append(f"row {r}: bad link {url!r}")
        continue
    jid = m.group(1)
    ids_in_sheet.append(jid)
    if jid in seen_ids:
        errors.append(f"row {r}: ALREADY-SEEN LEAK {jid} {ws.cell(r, 2).value}")
    if no != prev_no + 1:
        errors.append(f"row {r}: No. sequence broken ({no} after {prev_no})")
    prev_no = no
    if not (ws.cell(r, 5).value and len(str(ws.cell(r, 5).value)) > 100):
        warns.append(f"row {r}: short JD ({ws.cell(r, 2).value})")
    if ws.cell(r, 10).value != "Not Applied":
        errors.append(f"row {r}: status={ws.cell(r, 10).value}")

scores = [ws2.cell(r, 3).value for r in range(2, ws2.max_row + 1)]
if any(scores[i] < scores[i + 1] for i in range(len(scores) - 1)):
    errors.append("This-Week scores not descending")
if set(ids_in_sheet) != kept_ids:
    errors.append(f"sheet ids != kept set (sheet {len(set(ids_in_sheet))} vs kept {len(kept_ids)})")
if not (ws.cell(1, 1).value == "No." and ws.cell(1, 21).value == "Category"
        and ws2.cell(1, 13).value == "Application Link"):
    errors.append("header layout does not match the v3 spec")

print(f"verify_workbook: {wb_path}")
print(f"rows {len(ids_in_sheet)} | seen-ids checked {len(seen_ids)} | "
      f"scores {scores[0] if scores else '-'}..{scores[-1] if scores else '-'}")
print(f"ERRORS: {len(errors)}")
for e in errors[:15]:
    print("  E:", e)
print(f"WARNINGS: {len(warns)}")
for w in warns[:8]:
    print("  W:", w)
print("VERDICT:", "PASS" if not errors else "FAIL")
sys.exit(1 if errors else 0)

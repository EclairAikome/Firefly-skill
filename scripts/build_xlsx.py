"""Build the dated v3-style workbook (3 sheets, sorted by match score) and update master.

Sheets mirror SG_LinkedIn_Jobs_June2026_v3.xlsx (see references/v3_format.md):
  1. Job Applications (21 cols)  - sorted by score desc; No. follows that order
  2. This-Week by Fit (13 cols)  - same order; includes Why / Watch-outs
  3. Summary (3 cols)            - funnel, by-fit-tier, by-category, exclusions, methodology

Match scores come from scores.json (you, the agent) when present; otherwise a keyword
heuristic fallback so unattended runs still complete.
"""
import argparse, os, re, json, glob, datetime, sys
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
sys.path.insert(0, os.path.dirname(__file__))
import lib_common as L

SKILL_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MASTER = os.path.join(SKILL_DIR, "state", "master_jobs.jsonl")

ap = argparse.ArgumentParser()
ap.add_argument("--config", required=True)
ap.add_argument("--run", required=True)
a = ap.parse_args()
cfg = L.load_config(a.config)
run = a.run
today = datetime.date.today()

kept = L.read_json(os.path.join(run, "kept_candidates.json"))
scores = {}
sp = os.path.join(run, "scores.json")
if os.path.exists(sp):
    scores = L.read_json(sp)


def tier(score):
    return "High" if score >= 85 else ("Medium-High" if score >= 78 else "Medium")


recs = []
for k in kept:
    jid = k["jid"]
    s = scores.get(jid)
    if s:
        rec = dict(k)
        rec.update(fit=s.get("fit") or tier(int(s.get("score", 70))),
                   score=int(s.get("score", 70)), why=s.get("why", ""),
                   watch=s.get("watch", ""), category=s.get("category") or k["category_guess"],
                   industry=s.get("industry", ""))
    else:
        sc = int(k.get("heuristic_score", 65))
        rec = dict(k)
        rec.update(fit=tier(sc), score=sc,
                   why="Keyword overlap: " + ", ".join(k.get("kw_hits", [])[:6]),
                   watch="Auto-scored by keyword overlap; read the JD before applying.",
                   category=k["category_guess"], industry="")
    recs.append(rec)

recs.sort(key=lambda r: r["score"], reverse=True)
# Keep ALL jobs that passed the rule filter; do not truncate. The filter step's job is to
# remove disqualifiers, not to cap the count. Sorting is by fit score, highest first.

# ---------- styling ----------
HDR_FILL = PatternFill("solid", fgColor="FF1F4E78")
HDR_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFFFF")
DATA = Font(name="Arial", size=10)
LINK = Font(name="Arial", size=10, color="FF0563C1", underline="single")
TITLE_FONT = Font(name="Arial", size=13, bold=True, color="FF1F4E78")
SUBHDR = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
TOP = Alignment(vertical="top", wrap_text=True)
TOPN = Alignment(vertical="top", wrap_text=False)
CTR = Alignment(vertical="top", horizontal="center", wrap_text=True)
thin = Side(style="thin", color="FFD9D9D9")
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)


def header(ws, ncols, height=30):
    for c in range(1, ncols + 1):
        cell = ws.cell(1, c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORD
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = height


def widths(ws, w):
    for L_, v in w.items():
        ws.column_dimensions[L_].width = v


def set_date(ws, r, col, iso):
    try:
        y, m, d = [int(x) for x in iso.split("-")]
        ws.cell(r, col).value = datetime.date(y, m, d)
        ws.cell(r, col).number_format = "yyyy-mm-dd"
    except Exception:
        ws.cell(r, col).value = "TBD"


wb = Workbook()

# ===== Sheet 1: Job Applications =====
ws = wb.active
ws.title = "Job Applications"
H1 = ["No.", "Company Name", "Industry / Business Area", "Job Title", "Job Description (JD)",
      "Key Requirements", "Application Link", "Application Deadline", "Date Applied",
      "Application Status", "Interview Stage", "Contact Person", "Referral (Y/N)", "Priority",
      "Notes / Follow-up", "Company Size", "Size Tier", "This-Week Target", "Location",
      "Date Posted", "Category"]
ws.append(H1)
for i, r in enumerate(recs, 1):
    prio = "High" if r["score"] >= 85 else ("Medium" if r["score"] >= 75 else "Low")
    note = f'Posted {r["posted"]}; {r["work"]}; {r["location"]}; Fit {r["fit"]} ({r["score"]}); matched: {r["category"]}'
    ws.append([i, r["company"], r["industry"], r["title"], r["jd"], r["req"], "LinkedIn job link",
               "TBD", None, "Not Applied", None, None, None, prio, note, None, None, "YES",
               r["location"], None, r["category"]])
    rr = i + 1
    set_date(ws, rr, 20, r["posted"])
    lk = ws.cell(rr, 7); lk.hyperlink = r["link"]; lk.value = "LinkedIn job link"; lk.font = LINK
    for c in range(1, 22):
        cell = ws.cell(rr, c)
        if c != 7:
            cell.font = DATA
        cell.alignment = TOP if c in (3, 5, 6, 15, 19) else TOPN
        cell.border = BORD
    for c in (1, 14, 18):
        ws.cell(rr, c).alignment = CTR
header(ws, 21)
widths(ws, {"A": 5, "B": 24, "C": 22, "D": 30, "E": 60, "F": 46, "G": 32, "H": 13, "I": 11,
            "J": 16, "K": 15, "L": 13, "M": 11, "N": 9, "O": 34, "P": 15, "Q": 17, "R": 14,
            "S": 22, "T": 12, "U": 16})

# ===== Sheet 2: This-Week by Fit =====
ws2 = wb.create_sheet("This-Week by Fit")
H2 = ["Rank", "Fit", "Score", "Why it fits you", "Watch-outs", "Company Name", "Company Size",
      "Job Title", "Category", "Industry / Business Area", "Location", "Date Posted",
      "Application Link"]
ws2.append(H2)
for rank, r in enumerate(recs, 1):
    ws2.append([rank, r["fit"], r["score"], r["why"], r["watch"], r["company"], None, r["title"],
                r["category"], r["industry"], r["location"], None, "LinkedIn job link"])
    rr = ws2.max_row
    set_date(ws2, rr, 12, r["posted"])
    lk = ws2.cell(rr, 13); lk.hyperlink = r["link"]; lk.value = "LinkedIn job link"; lk.font = LINK
    for c in range(1, 14):
        cell = ws2.cell(rr, c)
        if c != 13:
            cell.font = DATA
        cell.alignment = TOP if c in (4, 5) else TOPN
        cell.border = BORD
    for c in (1, 2, 3):
        ws2.cell(rr, c).alignment = CTR
header(ws2, 13)
widths(ws2, {"A": 6, "B": 13, "C": 7, "D": 55, "E": 42, "F": 24, "G": 13, "H": 30, "I": 18,
             "J": 22, "K": 20, "L": 12, "M": 18})

# ===== Sheet 3: Summary =====
agg = L.read_json(os.path.join(run, "agg_stats.json")) if os.path.exists(os.path.join(run, "agg_stats.json")) else {}
dropped = L.read_json(os.path.join(run, "dropped.json")) if os.path.exists(os.path.join(run, "dropped.json")) else []
read_n = len(glob.glob(os.path.join(run, "details", "*.json")))
drop_reasons = Counter(re.split(r"[ (]", d["reason"])[0] for d in dropped)
fit_counts = Counter(r["fit"] for r in recs)
cat_counts = Counter(r["category"] for r in recs)

ws3 = wb.create_sheet("Summary")
def put(r, c, v, font=DATA, align=TOPN):
    cell = ws3.cell(r, c); cell.value = v; cell.font = font; cell.alignment = align; return cell

put(1, 1, f"LinkedIn Job Scout (auto-run {today.isoformat()}) - Summary", TITLE_FONT)
row = 3
put(row, 1, "Total relevant matches", DATA); put(row, 2, len(recs), DATA); row += 2

def table(title, headers, rows_):
    global row
    put(row, 1, title, Font(name="Arial", size=10, bold=True)); row += 1
    for ci, h in enumerate(headers, 1):
        cell = ws3.cell(row, ci); cell.value = h; cell.font = SUBHDR; cell.fill = HDR_FILL
        cell.alignment = TOPN; cell.border = BORD
    row += 1
    for rr in rows_:
        for ci, v in enumerate(rr, 1):
            cell = ws3.cell(row, ci); cell.value = v; cell.font = DATA; cell.border = BORD
        row += 1
    row += 1

table("Scrape funnel", ["Stage", "Count"], [
    ["Raw cards scraped", agg.get("raw_cards", "")],
    ["Unique jobs", agg.get("raw_unique", "")],
    ["Already applied/seen (excluded)", agg.get("excluded_applied", "")],
    ["Candidates after exclusion", agg.get("candidates", "")],
    ["Detail pages read", read_n],
    ["Final matches (this sheet)", len(recs)],
])
table("By fit tier", ["Fit Tier", "# Roles"], [[k, v] for k, v in fit_counts.most_common()])
table("By category", ["Category", "# Roles"], [[k, v] for k, v in cat_counts.most_common()])
table("Excluded after reading JD", ["Reason", "# Roles"], [[k, v] for k, v in drop_reasons.most_common()])

s = cfg.get("search", {})
notes = [
    f"Source: LinkedIn job search, {s.get('location_label','Singapore')} (geoId {s.get('geo_id','')}), "
    f"experience f_E={s.get('experience','')}, recency f_TPR={s.get('recency','')}.",
    f"Compiled: {today.isoformat()}. Queries: " + "; ".join(s.get("queries", [])) + ".",
    f"Excludes roles already applied/seen (master library + history sources).",
    f"Filters: Singapore base/business, minimum required experience < {cfg.get('filters',{}).get('max_years_exclusive',3)} years, "
    f"direct-sales/MLM roles dropped = {cfg.get('filters',{}).get('drop_direct_sales_mlm',True)}.",
    "Deadlines: LinkedIn rarely shows a hard deadline (TBD); check the link for 'No longer accepting'.",
    "De-duplicated by LinkedIn job id. Read-only scrape; this skill never auto-applies.",
]
for n in notes:
    put(row, 1, n, DATA, TOP); row += 1
widths(ws3, {"A": 70, "B": 14, "C": 22})
ws3.freeze_panes = "A2"

# ===== save + master append =====
outdir = cfg["output"]["dir"]
os.makedirs(outdir, exist_ok=True)
fname = f'{cfg["output"].get("basename","SG_LinkedIn_NewMatches")}_{today.isoformat()}.xlsx'
outpath = os.path.join(outdir, fname)
wb.save(outpath)

os.makedirs(os.path.dirname(MASTER), exist_ok=True)
with open(MASTER, "a", encoding="utf-8") as f:
    for r in recs:
        f.write(json.dumps({"jid": r["jid"], "company": r["company"], "title": r["title"],
                            "date": today.isoformat()}, ensure_ascii=False) + "\n")

print(f"SAVED {outpath}")
print(f"sheets: {wb.sheetnames}; rows: {len(recs)}; master appended: {len(recs)}")
